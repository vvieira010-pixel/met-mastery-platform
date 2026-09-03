import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
CURRENT = ROOT / 'current' / 'by-topic'
INCOMING = ROOT / 'incoming'
OUTPUT = ROOT / 'practice-studio-balance.xlsx'

def load_items(file):
    data = json.loads(file.read_text(encoding='utf-8'))
    if isinstance(data, list):
        return data
    return data.get('exercises') or data.get('questions') or data.get('prompts') or data.get('items') or []

def value(item, *keys):
    for key in keys:
        if item.get(key) not in (None, ''):
            return item[key]
    return ''

current_rows = []
topic_rows = []
skill_totals = {}
for skill_dir in sorted(CURRENT.iterdir()):
    if not skill_dir.is_dir():
        continue
    skill = skill_dir.name
    for file in sorted(skill_dir.glob('*.json')):
        if file.name == 'index.json':
            continue
        topic = file.stem
        items = load_items(file)
        topic_rows.append({'Skill': skill.title(), 'Topic': topic.replace('-', ' ').replace('_', ' ').title(), 'Exercises': len(items)})
        skill_totals[skill] = skill_totals.get(skill, 0) + len(items)
        for number, item in enumerate(items, 1):
            if not isinstance(item, dict):
                item = {'prompt': str(item)}
            options = value(item, 'options')
            if isinstance(options, dict):
                options = ' | '.join(f'{k}: {v}' for k, v in options.items())
            elif isinstance(options, list):
                options = ' | '.join(str(v) for v in options)
            current_rows.append({
                'Skill': skill.title(),
                'Topic': topic.replace('-', ' ').replace('_', ' ').title(),
                'Exercise #': number,
                'ID': value(item, 'id'),
                'Level': value(item, 'level'),
                'Type': value(item, 'type', 'category'),
                'Question / Prompt': value(item, 'question', 'prompt', 'instruction', 'template', 'intro'),
                'Options': options,
                'Correct Answer': value(item, 'correct_option', 'correct_answer', 'correct', 'answer'),
                'Time': value(item, 'seconds', 'time_minutes'),
                'Audio': value(item, 'audioSrc', 'audio'),
                'Image': value(item, 'image', 'imageSrc', 'image_url'),
            })

incoming_rows = []
for skill_dir in sorted(INCOMING.iterdir()):
    topic_dir = skill_dir / 'by-topic'
    if not topic_dir.is_dir():
        continue
    for file in sorted(topic_dir.glob('*.json')):
        if file.name == 'index.json':
            continue
        incoming_rows.append({'Skill': skill_dir.name.title(), 'Topic': file.stem.replace('-', ' ').replace('_', ' ').title(), 'Exercises': len(load_items(file)), 'Path': str(file.relative_to(ROOT))})

wb = Workbook()
summary = wb.active
summary.title = 'Balance'
summary.append(['Practice Studio Balance'])
summary.append(['This workbook covers the consolidated current collection. Incoming source packs are listed separately and are not double-counted.'])
summary.append([])
summary.append(['Skill', 'Exercises'])
for skill, count in sorted(skill_totals.items()):
    summary.append([skill.title(), count])
summary.append(['TOTAL', sum(skill_totals.values())])

def add_sheet(title, rows):
    ws = wb.create_sheet(title)
    if not rows:
        ws.append(['No records'])
        return ws
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, '') for h in headers])
    return ws

add_sheet('Topic Balance', sorted(topic_rows, key=lambda r: (r['Skill'], r['Topic'])))
add_sheet('Questions and Exercises', current_rows)
add_sheet('Incoming Inventory', incoming_rows)

for ws in wb.worksheets:
    ws.freeze_panes = 'A5' if ws.title == 'Balance' else 'A2'
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='1F4E78')
        cell.alignment = Alignment(vertical='center')
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_len = max(len(str(cell.value or '')) for cell in column_cells[:100])
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)
    ws.row_dimensions[1].height = 24

summary['A1'].font = Font(bold=True, size=16, color='FFFFFF')
summary.merge_cells('A1:B1')
summary.merge_cells('A2:B2')
summary['A2'].alignment = Alignment(wrap_text=True)
wb.save(OUTPUT)
print(f'Created {OUTPUT} with {sum(skill_totals.values())} current exercises.')
